"""
Copyright (C) 2015, MuChu Hsu
Contributed by Muchu Hsu (muchu1983@gmail.com)
This file is part of BSD license

<https://opensource.org/licenses/BSD-3-Clause>
"""
from openpyxl import Workbook
import os
from datetime import datetime
"""
excel 文件生成
"""
class XlwtWrapper:

    def __init__(self, filename=None):
        self.filename = filename
        self.wb = Workbook()
        self.ws = self.wb.active
        #加入標題
        self.ws["A1"] = "Date"
        self.ws["B1"] = "Entity"
        self.ws["C1"] = "Buy/Sell"
        self.ws["D1"] = "Name of Fund"
        self.ws["E1"] = "No. of Units"
        self.ws["F1"] = "Currency"
        self.ws["G1"] = "Unit Price"
        self.ws["H1"] = "Total Amount"
        self.ws["I1"] = "comment"
    
    #加入資料
    def addRowData(self, rowdata):
        self.ws.append(rowdata)
        
    #excel存檔
    def saveExcelFile(self):
        if self.filename == None:
            self.filename = datetime.now().strftime("%Y%m%d%H%M%S") + ".xlsx"
        self.wb.save(self.filename)